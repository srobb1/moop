#!/usr/bin/env python3
"""
update_evidence_sheet_paths.py — fix TRACK_PATH in the multi-sheet track xlsx.

Two problems to fix in each organism sheet:

1. GENESET MISSING: TRACK_PATH has /{Assembly}/gff/{file} — the geneset dir is absent.
   Fix: /{Assembly}/gff/{file} → /{Assembly}/{Geneset}/gff/{file}

2. WRONG EVIDENCE FILENAME: evidence rows (repeatmasker, repeatrunner, est2genome*,
   protein2genome, blastx) all point to ACA1.putative_function.gff instead of
   their real evidence GFF.
   Fix: gff/{wrong} → gff/evidence/{correct} (using evidence.txt as source of truth)

Usage:
  python3 scripts/update_evidence_sheet_paths.py \\
      --input  organisms/Amphimedon_queenslandica/.../track_sheets_reformatted.xlsx \\
      --output track_sheets_updated.xlsx \\
      --evidence evidence.txt \\
      [--tracks-listing tracks_files.txt]

Requires: openpyxl
"""

import argparse
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

TRACKS_URL_BASE = 'https://tracks.stowers.org:8080/moop/data/tracks/'

# Mapping from TRACKS organism name → GENOMES organism name (for evidence.txt lookup)
TRACKS_TO_GENOMES = {
    'Molossus_molossus':      'Mollosus_mollosus',
    'Murina_feae':            'Murina_aurata_feae',
    'Musonycteris_harrisoni': 'Musonycteris_harrisonii',
}

# Known assembly → geneset for all bat organisms (from copy_evidence_to_tracks.py MAPPING)
# Key: (tracks_organism, assembly)  Value: geneset
BAT_GENESET = {
    ('Anoura_caudifer',             'GCA_004027475.1'):               'SIMR_2025-01-24',
    ('Antrozous_pallidus',          'GCA_007922775.1'):               'SIMR_2025-01-24',
    ('Artibeus_jamaicensis',        'GCF_021234435.1'):               'RS_2023_03',
    ('Carollia_perspicillata',      'GCA_004027735.1'):               'SIMR_2025-01-24',
    ('Craseonycteris_thonglongyai', 'GCA_004027555.1'):               'SIMR_2025-01-24',
    ('Cynopterus_brachyotis',       'GCA_009793145.1'):               'SIMR_2025-01-24',
    ('Desmodus_rotundus',           'GCF_022682495.1'):               'RS_2023_02',
    ('Eidolon_dupreanum',           'ASM46528v1'):                    'SIMR_2025-01-24',
    ('Eidolon_helvum',              'GCA_000465285.1'):               'SIMR_2025-01-24',
    ('Eonycteris_spelaea',          'GCA_003508835.1'):               'SIMR_2025-01-24',
    ('Eptesicus_fuscus',            'GCF_027574615.1'):               'RS_2023_03',
    ('Hipposideros_armiger',        'GCF_001890085.2'):               'RS_100',
    ('Hipposideros_galeritus',      'GCA_004027415.1'):               'SIMR_2025-01-24',
    ('Lasiurus_borealis',           'GCA_004026805.1'):               'SIMR_2025-01-24',
    ('Lasiurus_cinereus',           'GCA_011751095.1'):               'SIMR_2025-01-23',
    ('Leptonycteris_nivalis',       'Lnivalis_consensus_genome'):     'SIMR_2025-01-24',
    ('Leptonycteris_yerbabuenae',   'Lyerbabuenae_genome'):           'SIMR_2025-01-24',
    ('Macroglossus_sobrinus',       'GCA_004027375.1'):               'SIMR_2025-01-24',
    ('Macrotus_californicus',       'GCA_007922815.1'):               'SIMR_2025-01-24',
    ('Macrotus_waterhousii',        'Mwaterhousii_consensus_genome'): 'SIMR_2025-01-24',
    ('Megaderma_lyra',              'GCA_004026885.1'):               'SIMR_2025-01-24',
    ('Micronycteris_hirsuta',       'GCA_004026765.1'):               'SIMR_2025-01-24',
    ('Miniopterus_natalensis',      'GCF_001595765.1'):               'RS_100',
    ('Miniopterus_schreibersii',    'GCA_004026525.1'):               'SIMR_2025-01-24',
    ('Molossus_molossus',           'GCF_014108415.1'):               'RS_100',
    ('Mormoops_blainvillei',        'GCA_004026545.1'):               'SIMR_2025-01-24',
    ('Murina_feae',                 'GCA_004026665.1'):               'SIMR_2025-01-24',
    ('Musonycteris_harrisoni',      'Mharrisoni_consensus_genome'):   'SIMR_2025-01-24',
    ('Myotis_brandtii',             'GCF_000412655.1'):               'RS_101',
    ('Myotis_davidii',              'GCF_000327345.1'):               'RS_101',
    ('Myotis_lucifugus',            'GCF_000147115.1'):               'RS_102',
    ('Myotis_myotis',               'GCF_014108235.1'):               'RS_100',
    ('Myotis_septentrionalis',      'myse_ont_racon_pilon_HiC'):      'SIMR_2025-01-24',
    ('Noctilio_leporinus',          'GCA_004026585.1'):               'SIMR_2025-01-24',
    ('Nycticeius_humeralis',        'GCA_007922795.1'):               'SIMR_2025-01-24',
    ('Phyllostomus_discolor',       'GCF_004126475.2'):               'RS_101',
    ('Phyllostomus_hastatus',       'GCF_019186645.2'):               'RS_100',
    ('Pipistrellus_kuhlii',         'GCF_014108245.1'):               'RS_101',
    ('Pipistrellus_pipistrellus',   'GCA_004026625.1'):               'SIMR_2025-01-24',
    ('Pteropus_alecto',             'GCF_000325575.1'):               'RS_102',
    ('Pteropus_rufus',              'Pteropus_rufus_HiC'):            'SIMR_2025-01-24',
    ('Pteropus_vampyrus',           'GCF_000151845.1'):               'RS_101',
    ('Rhinolophus_ferrumequinum',   'GCF_004115265.2'):               'RS_2023_02',
    ('Rousettus_aegyptiacus',       'GCF_014176215.1'):               'Release-101',
    ('Rousettus_madagascariensis',  'GCA_028533395.1'):               'SIMR_2025-01-24',
    ('Sturnira_hondurensis',        'GCF_014824575.3'):               'Release-100.20210611',
    ('Tadarida_brasiliensis',       'GCA_004025005.1'):               'SIMR_2025-01-24',
    ('Tonatia_saurophila',          'GCA_004024845.1'):               'SIMR_2025-01-24',
}

EVIDENCE_TYPES = ['repeatmasker', 'repeatrunner', 'protein2genome', 'blastx']


def load_tracks_geneset_map(tracks_listing):
    """
    Parse tracks_files.txt to build (org, assembly) → geneset for non-bat organisms.
    Complements BAT_GENESET for organisms not in the hardcoded table.
    """
    import re as _re
    DATA_DIRS = {'RNAseq','RNASEQ','RNA_Seq','whole_genome_alignment','chipSeq',
                 'atac-seq','DNA_Seq','ISO_Seq','RAD_Seq','methylation','gff'}
    TRACKS_BASE = '/var/www/privatehtml/moop/data/tracks'

    result = {}
    tree = defaultdict(lambda: defaultdict(set))
    with open(tracks_listing) as fh:
        for line in fh:
            line = line.strip()
            if not line.startswith(TRACKS_BASE + '/'):
                continue
            rel = line[len(TRACKS_BASE)+1:]
            parts = rel.split('/')
            if len(parts) >= 3:
                tree[parts[0]][parts[1]].add(parts[2])

    for org, asms in tree.items():
        for asm, subdirs in asms.items():
            if (org, asm) in BAT_GENESET:
                continue  # already covered
            genesets = [d for d in subdirs if d not in DATA_DIRS and not _re.match(r'^\.', d)]
            if len(genesets) == 1:
                result[(org, asm)] = genesets[0]
    return result


def load_evidence_map(evidence_file):
    """Returns: genomes_org → { type_key: filename }"""
    ev = defaultdict(dict)
    with open(evidence_file) as fh:
        for line in fh:
            path = line.strip()
            if not path.endswith('.gff'):
                continue
            parts = path.split('/')
            try:
                bat_idx = parts.index('bat')
                org = parts[bat_idx + 1]
            except (ValueError, IndexError):
                continue
            fname = parts[-1]
            for t in EVIDENCE_TYPES:
                if fname.startswith(t + '.'):
                    ev[org][t] = fname
                    break
            else:
                m = re.match(r'^(est2genome:[^.]+)\.', fname)
                if m:
                    ev[org][m.group(1)] = fname
    return ev


def track_id_to_type_key(track_id):
    """Return evidence type key or None if not an evidence row."""
    tid = track_id.lower()
    for t in EVIDENCE_TYPES:
        if t in tid:
            return t
    m = re.match(r'^est2genome_(.+)$', track_id, re.IGNORECASE)
    if m:
        return 'est2genome:' + m.group(1)
    return None


def needs_geneset_inserted(url, assembly):
    """True if the URL has /{assembly}/gff/ without a geneset directory between them."""
    # Pattern: /tracks/{org}/{asm}/gff/  (gff directly after assembly)
    asm_esc = re.escape(assembly)
    return bool(re.search(rf'/{asm_esc}/gff/', url))


def insert_geneset(url, assembly, geneset):
    """Replace /{assembly}/gff/ with /{assembly}/{geneset}/gff/ in URL."""
    asm_esc = re.escape(assembly)
    return re.sub(rf'(/{asm_esc})/gff/', rf'\1/{geneset}/gff/', url, count=1)


def process_workbook(input_path, output_path, ev_map, extra_geneset_map):
    import openpyxl
    wb = openpyxl.load_workbook(input_path)

    geneset_map = {**BAT_GENESET, **extra_geneset_map}

    total_geneset_fixes = 0
    total_evidence_fixes = 0
    skipped_sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == 'Table of Contents':
            continue

        # Parse organism and assembly from sheet name: {Org}_{Assembly}
        # Assembly accessions contain dots/underscores — split on first underscore run
        # that looks like an accession pattern
        m = re.match(
            r'^(.+?)_((?:GC[AF]_\d+\.\d+|ASM\d+v\d+|[A-Z][a-z]+_(?:consensus_genome|HiC|ont_.+)|.+_\d{4}-\d{2}|myse_ont_.+|Lnivalis_consensus_genome|Lyerbabuenae_genome|Mwaterhousii_consensus_genome|Mharrisoni_consensus_genome|Pteropus_rufus_HiC|Scal100|schMedS3h1|GCA_\d+\.\d+|LPT|CCA1|Mcap_\d+|[A-Z]+\d+))$',
            sheet_name
        )
        if not m:
            # Simpler split: last underscore-delimited segment that looks like accession
            parts = sheet_name.rsplit('_', maxsplit=1)
            if len(parts) == 2:
                org, asm = parts
            else:
                print(f'  SKIP (cannot parse sheet name): {sheet_name}', file=sys.stderr)
                skipped_sheets.append(sheet_name)
                continue
        else:
            org, asm = m.group(1), m.group(2)

        geneset = geneset_map.get((org, asm))
        genomes_org = TRACKS_TO_GENOMES.get(org, org)

        # Find header row and column indices
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h or '').strip().lower() for h in header_row]
        try:
            col_tid   = headers.index('track_id') + 1   # 1-based for openpyxl
            col_fname = headers.index('filename')   + 1
            col_tpath = headers.index('track_path') + 1
        except ValueError:
            print(f'  SKIP (missing required header): {sheet_name}', file=sys.stderr)
            skipped_sheets.append(sheet_name)
            continue

        org_ev = ev_map.get(genomes_org, {})
        sheet_geneset_fixes = 0
        sheet_evidence_fixes = 0

        for row in ws.iter_rows(min_row=2):
            track_id   = str(row[col_tid - 1].value   or '').strip()
            filename   = str(row[col_fname - 1].value or '').strip()
            track_path = str(row[col_tpath - 1].value or '').strip()

            if not track_id or not track_path:
                continue

            new_path = track_path
            new_fname = filename

            # Step 1: insert geneset if missing
            if geneset and needs_geneset_inserted(track_path, asm):
                new_path = insert_geneset(new_path, asm, geneset)
                sheet_geneset_fixes += 1

            # Step 1b: fill in empty TRACK_PATH from FILENAME (relative path → full URL)
            # Applies to rows like RNAseq/isoseq/ISOseq.bw where TRACK_PATH was left blank.
            if not new_path and new_fname and '/' in new_fname:
                # Use assembly-level base URL (no geneset — these files live directly under {assembly}/)
                base = f'{TRACKS_URL_BASE}{org}/{asm}/{new_fname}'
                new_path = base

            # Step 2: fix evidence filename
            type_key = track_id_to_type_key(track_id)
            if type_key and org_ev:
                correct_fname = org_ev.get(type_key)
                if not correct_fname:
                    # case-insensitive match for est2genome species names
                    if type_key.startswith('est2genome:'):
                        for k, v in org_ev.items():
                            if k.lower() == type_key.lower():
                                correct_fname = v
                                break

                if correct_fname:
                    new_evidence = f'evidence/{correct_fname}'
                    # Replace the filename portion after /gff/
                    new_path = re.sub(r'/gff/[^/]+$', f'/gff/{new_evidence}', new_path)
                    new_fname = new_evidence
                    sheet_evidence_fixes += 1

            if new_path != track_path:
                row[col_tpath - 1].value = new_path
            if new_fname != filename:
                row[col_fname - 1].value = new_fname

        if sheet_geneset_fixes or sheet_evidence_fixes:
            print(f'  {sheet_name}: +geneset={sheet_geneset_fixes}, evidence_fixes={sheet_evidence_fixes}')
        total_geneset_fixes += sheet_geneset_fixes
        total_evidence_fixes += sheet_evidence_fixes

    wb.save(output_path)
    print(f'\nTotal: geneset_insertions={total_geneset_fixes}, evidence_fixes={total_evidence_fixes}')
    if skipped_sheets:
        print(f'Skipped sheets: {len(skipped_sheets)}')
        for s in skipped_sheets:
            print(f'  {s}')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input',    required=True)
    ap.add_argument('--output',   required=True)
    ap.add_argument('--evidence', default='evidence.txt')
    ap.add_argument('--tracks-listing', default='tracks_files.txt',
                    help='tracks_files.txt for non-bat geneset lookup')
    args = ap.parse_args()

    print(f'Loading evidence map...', file=sys.stderr)
    ev_map = load_evidence_map(args.evidence)
    print(f'  {len(ev_map)} organisms with evidence files', file=sys.stderr)

    extra = {}
    if os.path.exists(args.tracks_listing):
        extra = load_tracks_geneset_map(args.tracks_listing)
        print(f'  {len(extra)} extra (non-bat) geneset entries from tracks listing', file=sys.stderr)

    print(f'\nProcessing {args.input}...')
    process_workbook(args.input, args.output, ev_map, extra)
    print(f'\nWrote {args.output}')


if __name__ == '__main__':
    main()
