#!/usr/bin/env python3
"""
Parse metadata.contents.txt and scp_commands.sh to build reformatted
track sheets as a multi-tab XLSX (one sheet per organism/assembly).

New column format matches nvec_tracks_updated.tsv:
  TRACK_ID, NAME, technique, CATEGORY, institute, source, experiment,
  developmental-stage, tissue, condition, summary, citation, project,
  accession, date, analyst, SCIPRJ, ACCESS, biosample, NGS_file, MLONG,
  FILENAME, TRACK_PATH, ASSEMBLY_1, ASSEMBLY_2, BED1_PATH, BED2_PATH, MAF

TRACK_PATH is filled from the scp_commands.sh destination paths.
"""

import re
import csv
import io
import os
import sys
import json
import sqlite3
from collections import defaultdict

# --- try to import openpyxl ---
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: python3 -m pip install --user openpyxl")
    sys.exit(1)

# ──────────────────────────────────────────────────────────────
# New column order (matches nvec_tracks_updated.tsv)
# ──────────────────────────────────────────────────────────────
TRACKS_BASE_URL = 'https://tracks.stowers.org:8080/moop/data/tracks/'
TRACKS_FS_PREFIX = '/var/www/privatehtml/moop/data/tracks/'

NEW_HEADERS = [
    "TRACK_ID", "NAME", "technique", "CATEGORY", "institute", "source",
    "experiment", "developmental-stage", "tissue", "condition", "summary",
    "citation", "project", "accession", "date", "analyst", "SCIPRJ",
    "ACCESS", "biosample", "NGS_file", "MLONG", "FILENAME", "TRACK_PATH",
    "ASSEMBLY_1", "ASSEMBLY_2", "BED1_PATH", "BED2_PATH", "MAF",
]

# Old CSV header → new header (for columns that just rename)
OLD_TO_NEW = {
    "label":               "TRACK_ID",
    "key":                 "NAME",
    "technique":           "technique",
    "category":            "CATEGORY",
    "institute":           "institute",
    "source":              "source",
    "experiment":          "experiment",
    "developmental-stage": "developmental-stage",
    "condition":           "condition",
    "summary":             "summary",
    "citation":            "citation",
    "accession":           "accession",
    "date":                "date",
    "analyst":             "analyst",
    "project":             "SCIPRJ",    # old "project" = SCI number → SCIPRJ
    "filename":            "FILENAME",
    "private tracks":      "ACCESS",    # old column 17
}

ACCESS_VALUE_MAP = {
    'private':      'COLLABORATOR',
    'restricted':   'COLLABORATOR',
    'login':        'COLLABORATOR',
    'logged_in':    'COLLABORATOR',
    'collaborator': 'COLLABORATOR',
    'public':       'PUBLIC',
    'ip_in_range':  'IP_IN_RANGE',
    'ip':           'IP_IN_RANGE',
    'admin':        'ADMIN',
}

def normalize_access(val):
    return ACCESS_VALUE_MAP.get(val.strip().lower(), val.strip())


# ──────────────────────────────────────────────────────────────
# Parse tracks_files.txt — full file listing of the tracks server
# ──────────────────────────────────────────────────────────────
def parse_tracks_files(filepath):
    """
    Parse a find-output file listing of /var/www/privatehtml/moop/data/tracks/.
    Returns: {(organism, assembly) -> {relative_path -> url}}
    relative_path is everything after organism/assembly/ (e.g. 'RNAseq/foo.bw' or 'gff/bar.gff')
    Also keyed by bare filename for convenience.
    """
    tracks_map = defaultdict(dict)
    if not os.path.exists(filepath):
        return tracks_map

    prefix = '/var/www/privatehtml/moop/data/tracks/'

    with open(filepath) as fh:
        for line in fh:
            path = line.strip()
            if not path.startswith(prefix):
                continue
            rel = path[len(prefix):]
            parts = rel.split('/')
            if len(parts) < 3:
                continue   # directory-level entry, skip
            organism = parts[0]
            assembly = parts[1]
            rel_path = '/'.join(parts[2:])   # e.g. 'RNAseq/MOLNG-2707/foo.bw'
            bare     = parts[-1]             # e.g. 'foo.bw'
            url = path.replace('/var/www/privatehtml/moop/data/tracks/', TRACKS_BASE_URL, 1)
            key = (organism, assembly)
            tracks_map[key][rel_path] = url
            # Also index by bare filename (may collide — last one wins, acceptable)
            tracks_map[key][bare]     = url

    return tracks_map


# ──────────────────────────────────────────────────────────────
# Parse genomes_find_results.txt → bare_filename → expected tracks URL
# ──────────────────────────────────────────────────────────────
def parse_genomes_find_results(filepath, org_map, geneset_map):
    """
    Build {bare_filename -> tracks_url} from the output of find_missing_tracks.sh.
    Uses org_map (from scp_commands) to derive the tracks destination dir per organism.
    Returns {bare_filename -> url}
    """
    fn_to_url = {}
    if not os.path.exists(filepath):
        return fn_to_url

    # Section header → organism label (strip '=== ... ===')
    current_org_label = ''
    found = {}   # bare -> (org_label, genomes_path) — prefer wei/ if duplicate

    # Org label → list of (code, organism, assembly, dest_dir)
    from collections import defaultdict as _dd
    label_to_info = _dd(list)
    for code, info in org_map.items():
        label_to_info[info['organism']].append((code, info['organism'], info['assembly'], info['dest_dir']))

    with open(filepath) as fh:
        for line in fh:
            line = line.strip()
            if line.startswith('==='):
                current_org_label = line.strip('= ')
            elif line.startswith('/var/other_data/'):
                bare = os.path.basename(line)
                if bare not in found or '/wei/' in line:
                    found[bare] = (current_org_label, line)

    # NV2g.20240221.gff — mirror path of NV2g.20231129.gff already on tracks
    fn_to_url['NV2g.20240221.gff'] = (
        TRACKS_BASE_URL + 'Nematostella_vectensis/GCA_033964005.1/'
        'NV2/aligned/tcs_v2/20240221/NV2g.20240221.gff'
    )

    # Danio rerio WGA — metadata CSV has _on_nf.bb but actual files on tracks are _on_dr.bb
    _danio = TRACKS_BASE_URL + 'Danio_rerio/GCF_000002035.5/whole_genome_alignment/'
    fn_to_url['fish_projected_on_nf.bb'] = _danio + 'fish_projected_on_dr.bb'
    fn_to_url['all_projected_on_nf.bb']  = _danio + 'all_projected_on_dr.bb'

    # Chamaeleo CCA1 conservation — metadata has .summary.bb but actual file has no .summary
    _cca1 = TRACKS_BASE_URL + 'Chamaeleo_calyptratus/CCA1/conservation/'
    fn_to_url['rattlesnake_mytree.summary.bb'] = _cca1 + 'rattlesnake_mytree.bb'
    fn_to_url['chameleon_mytree.summary.bb']   = _cca1 + 'chameleon_mytree.bb'

    for bare, (org_label, genomes_path) in found.items():
        if bare in fn_to_url:
            continue
        entries = label_to_info.get(org_label, [])
        if not entries:
            continue

        # Pick the right entry when multiple assemblies exist for an organism
        chosen = entries[0]
        if len(entries) > 1:
            for entry in entries:
                code = entry[0]
                if 'ETRf' in code and 'ETRf' in genomes_path: chosen = entry; break
                if 'ETRm' in code and 'ETRm' in genomes_path: chosen = entry; break
                if 'v31' in code and 'v3.1' in genomes_path:  chosen = entry; break
                if 'schMedS3' in code and 'schMedS3' in genomes_path: chosen = entry; break

        code, organism, assembly, dest_dir = chosen
        gs = geneset_map.get((organism, assembly), '')
        if gs:
            url = TRACKS_BASE_URL + organism + '/' + assembly + '/' + gs + '/gff/' + bare
        else:
            url = TRACKS_BASE_URL + organism + '/' + assembly + '/gff/' + bare
        fn_to_url[bare] = url

    return fn_to_url


# ──────────────────────────────────────────────────────────────
# Parse scp_commands.sh
# ──────────────────────────────────────────────────────────────
def parse_scp_commands(filepath):
    """
    Returns:
      org_map  : {code -> {'organism': str, 'assembly': str, 'dest_dir': str}}
      file_map : {code -> {filename -> full_dest_path}}
    """
    org_map = {}
    file_map = defaultdict(dict)
    current_code = None

    with open(filepath, 'r') as f:
        for line in f:
            line = line.rstrip('\n')

            # Comment: # ACA1_v1 → /var/www/.../Organism/Assembly/gff
            m = re.match(
                r'^# (\S+) → (/var/www/privatehtml/moop/data/tracks/([^/]+)/([^/]+)/gff)',
                line
            )
            if m:
                current_code = m.group(1)
                dest_dir      = m.group(2)
                organism      = m.group(3)
                assembly      = m.group(4)
                org_map[current_code] = {
                    'organism': organism,
                    'assembly': assembly,
                    'dest_dir': dest_dir,
                }
                continue

            # Reset current_code on TRACKS_DEST (unmapped) comment
            m2 = re.match(r'^# (\S+) → TRACKS_DEST', line)
            if m2:
                current_code = None
                continue

            # Active scp line: scp src tracks:/dest/path/file.gff [# comment]
            m3 = re.match(r'^scp .+ tracks:(/[^\s#]+)', line)
            if m3 and current_code:
                dest_path = m3.group(1).rstrip()
                filename  = os.path.basename(dest_path)
                # Convert filesystem path to URL
                url = dest_path.replace(TRACKS_FS_PREFIX, TRACKS_BASE_URL, 1)
                file_map[current_code][filename] = url

    return org_map, file_map


# ──────────────────────────────────────────────────────────────
# Parse scp_evidence.sh for evidence file map + geneset map
# ──────────────────────────────────────────────────────────────
def parse_scp_evidence(filepath):
    """
    Returns:
      evidence_map : {short_code -> {filename -> url}}
      geneset_map  : {(organism, assembly) -> geneset}

    short_code is the directory name before 'evidence/' in the genomes source path,
    e.g. /var/other_data/organisms/bat/Anoura_caudifer/ACA1/evidence/file.gff → 'ACA1'
    """
    evidence_map = defaultdict(dict)
    geneset_map  = {}
    if not os.path.exists(filepath):
        return evidence_map, geneset_map

    mkdir_re = re.compile(
        r'^ssh tracks mkdir -p '
        r'/var/www/privatehtml/moop/data/tracks/([^/]+)/([^/]+)/([^/]+)/gff/evidence'
    )
    scp_re = re.compile(r'^scp genomes:(/[^\s]+)\s+tracks:(/[^\s]+)')

    with open(filepath) as fh:
        for line in fh:
            line = line.rstrip()
            m = mkdir_re.match(line)
            if m:
                geneset_map[(m.group(1), m.group(2))] = m.group(3)
                continue
            m = scp_re.match(line)
            if m:
                src, dest = m.group(1), m.group(2)
                parts = src.split('/')
                try:
                    ev_idx     = parts.index('evidence')
                    short_code = parts[ev_idx - 1]
                    filename   = parts[ev_idx + 1]
                except (ValueError, IndexError):
                    continue
                url = dest.replace(TRACKS_FS_PREFIX, TRACKS_BASE_URL, 1)
                evidence_map[short_code][filename] = url

    return evidence_map, geneset_map


def supplement_geneset_map(geneset_map, mv_script):
    """
    Enrich geneset_map from move_gff_to_geneset.sh, which covers all bat organisms
    including those with no evidence files (e.g. Lasiurus_cinereus).
    Format: mkdir -p /var/.../tracks/{Org}/{Asm}/{Geneset}/gff
    """
    if not os.path.exists(mv_script):
        return
    mkdir_re = re.compile(
        r'mkdir -p /var/www/privatehtml/moop/data/tracks/([^/]+)/([^/]+)/([^/]+)/gff\b'
    )
    with open(mv_script) as fh:
        for line in fh:
            m = mkdir_re.search(line)
            if m:
                key = (m.group(1), m.group(2))
                if key not in geneset_map:
                    geneset_map[key] = m.group(3)


# ──────────────────────────────────────────────────────────────
# Evidence TRACK_ID → evidence file type prefix
# ──────────────────────────────────────────────────────────────
_EVIDENCE_TRACK_IDS = {
    'repeatmasker_maker': 'repeatmasker',
    'repeatmasker':       'repeatmasker',
    'repeatrunner_maker': 'repeatrunner',
    'repeatrunner':       'repeatrunner',
    'protein2genome':     'protein2genome',
    'blastx':             'blastx',
}

# Some organisms have a different directory code on the genomes server than in
# their metadata CSV (e.g. Pteropus_vampyrus metadata uses PVA1 but evidence
# files are under PVA2).
_CODE_REMAP = {
    'PVA1': 'PVA2',
}

def _track_id_to_evidence_type(track_id):
    """
    Map a TRACK_ID to its evidence filename prefix, or None if not evidence.
      'repeatmasker_maker'           → 'repeatmasker'
      'est2genome_Desmodus_rotundus' → 'est2genome:Desmodus_rotundus'
      'aca1_protein2genome'          → 'protein2genome'
      'aca1_blastx'                  → 'blastx'
    """
    t = track_id.lower()
    if t in _EVIDENCE_TRACK_IDS:
        return _EVIDENCE_TRACK_IDS[t]
    if t.startswith('est2genome_'):
        return 'est2genome:' + track_id[len('est2genome_'):]
    if 'protein2genome' in t:
        return 'protein2genome'
    if 'blastx' in t:
        return 'blastx'
    return None


# ──────────────────────────────────────────────────────────────
# Parse metadata.contents.txt
# ──────────────────────────────────────────────────────────────
def parse_metadata_contents(filepath):
    """
    Parses blocks delimited by:
        ######
        /path/to/CODE/includes/xxx_metadata.csv
        ######
        label,key,...   ← header
        row...
        #### END ######

    Returns {code -> [list of row dicts keyed by old header names]}
    """
    metadata = defaultdict(list)
    current_code = None
    header = None
    in_data = False   # True after the second "######" (CSV content started)
    path_seen = False # True after we've seen the path line

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\n')

            if line.startswith('#### END'):
                # End of block
                current_code = None
                header = None
                in_data = False
                path_seen = False
                continue

            if line == '######':
                if not path_seen:
                    # First delimiter — next line will be the path
                    path_seen = False  # wait for path
                    in_data = False
                else:
                    # Second delimiter — CSV data starts next
                    in_data = True
                continue

            # Path line (between the two ######)
            if not in_data and not path_seen:
                m = re.search(r'/data/([^/]+)/includes/', line)
                if m:
                    current_code = m.group(1)
                    path_seen = True
                continue

            # CSV data lines
            if in_data and current_code:
                if not line.strip():
                    continue
                try:
                    reader = csv.reader(io.StringIO(line))
                    row = next(reader)
                except Exception:
                    continue

                if all(c.strip() == '' for c in row):
                    continue

                if header is None:
                    # First non-empty line is the header
                    header = [c.strip().lower() for c in row]
                    continue

                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(header) and header[i]:
                        row_dict[header[i]] = val.strip()
                metadata[current_code].append(row_dict)

    return metadata


# ──────────────────────────────────────────────────────────────
# Convert one old row dict → new row dict
# ──────────────────────────────────────────────────────────────
def convert_row(old_row, file_path_map, evidence_map=None, geneset='', tracks_map=None, genomes_map=None):
    """
    file_path_map : {filename -> url} for this organism code (from scp_commands.sh)
    evidence_map  : {filename -> url} for this organism's evidence files (from scp_evidence.sh)
    geneset       : geneset directory name to insert into scp_commands-derived URLs
    tracks_map    : {rel_path_or_bare -> url} for this (organism, assembly) from tracks_files.txt
    """
    new = {col: '' for col in NEW_HEADERS}

    for old_col, new_col in OLD_TO_NEW.items():
        val = old_row.get(old_col, '')
        if val:
            if new_col == 'ACCESS':
                new[new_col] = normalize_access(val)
            else:
                new[new_col] = val

    track_id = new.get('TRACK_ID', '')
    ev_type  = _track_id_to_evidence_type(track_id)

    # For evidence-type TRACK_IDs, derive URL from evidence_map via type prefix.
    # Exceptions — skip evidence lookup and fall through to filename-based lookup if:
    #   • FILENAME contains '/' (explicit subdirectory path, e.g. MAKER_082015/file.gff)
    #   • FILENAME ends with '.evidence.gff' (combined MAKER evidence GFF like pmz.evidence.gff)
    filename_raw = new.get('FILENAME', '').strip()
    bare_fn = os.path.basename(filename_raw)
    _is_combined_evidence = bare_fn.lower().endswith('.evidence.gff') or bare_fn.lower().endswith('.evidence.gff3')
    if ev_type is not None and '/' not in filename_raw and not _is_combined_evidence:
        prefix = ev_type.lower() + '.'
        # Try scp_evidence.sh map first
        if evidence_map:
            for fn, url in sorted(evidence_map.items()):
                if fn.lower().startswith(prefix):
                    new['TRACK_PATH'] = url
                    new['FILENAME']   = fn
                    return new
        # Fallback: full tracks server listing
        if tracks_map:
            for fn, url in sorted(tracks_map.items()):
                bare = os.path.basename(fn)
                if bare.lower().startswith(prefix):
                    new['TRACK_PATH'] = url
                    new['FILENAME']   = bare
                    return new
        # Prefix lookup found nothing — fall through to FILENAME-based lookup.
        # Handles organisms like Lasiurus_cinereus where evidence tracks share a
        # combined putative_function.gff instead of having per-type evidence files.

    # Fill TRACK_PATH from scp_commands file map
    filename = new.get('FILENAME', '').strip()
    if filename:
        dest = file_path_map.get(filename, '')
        if not dest:
            bare = os.path.basename(filename)
            dest = file_path_map.get(bare, '')
        if dest and geneset:
            # scp_commands.sh paths lack geneset; insert it before /gff/
            dest = re.sub(r'/gff(/|$)', f'/{geneset}/gff\\1', dest, count=1)
        if dest:
            new['TRACK_PATH'] = dest
            return new

    # Fallback: look up in full tracks server file listing
    if filename and tracks_map:
        dest = tracks_map.get(filename, '')
        if not dest:
            bare = os.path.basename(filename)
            dest = tracks_map.get(bare, '')
        if dest:
            new['TRACK_PATH'] = dest
            return new

    # Fallback: genomes find results (files located on genomes server → expected tracks URL)
    if filename and genomes_map:
        bare = os.path.basename(filename)
        dest = genomes_map.get(filename, '') or genomes_map.get(bare, '')
        new['TRACK_PATH'] = dest

    return new


# ──────────────────────────────────────────────────────────────
# XLSX styling helpers
# ──────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="0891B2")   # teal accent
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(size=9)
FREEZE_ROW    = 2   # freeze row 1 (header)

def style_header_row(ws, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

def auto_width(ws, headers):
    for i, h in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        # Width: header text + 4, capped at 40
        ws.column_dimensions[col_letter].width = min(len(h) + 4, 40)


# ──────────────────────────────────────────────────────────────
# Groups + common name lookups
# ──────────────────────────────────────────────────────────────
def load_groups(base):
    """Return {(organism, assembly) -> 'Group1, Group2'} from organism_assembly_groups.json."""
    path = os.path.join(base, 'metadata', 'organism_assembly_groups.json')
    if not os.path.exists(path):
        return {}
    with open(path, 'r') as f:
        data = json.load(f)
    return {
        (e['organism'], e['assembly']): ', '.join(e.get('groups', []))
        for e in data
    }


# Filenames that are genome sequences or archives — not track files.
# Rows whose FILENAME (bare) matches these are dropped from the output sheets.
SKIP_TRACK_FILENAMES = {
    'evidence.tar.gz',          # Chamaeleo: archive, individual evidence GFFs are the real tracks
    'SmedSxl_genome_v3.1.nt',  # Schmidtea: nucleotide sequence, not a track
}
SKIP_TRACK_EXTENSIONS = {'.tar.gz', '.tar', '.zip'}
# Genome FASTA extensions — only skip when TRACK_ID is 'DNA' (reference sequence row)
GENOME_FASTA_EXTENSIONS = {'.fa', '.fna', '.fasta', '.fas'}


def should_skip_row(old_row):
    """Return True if this metadata row should be excluded from the output sheet."""
    fn    = old_row.get('filename', '').strip()
    label = old_row.get('label', '').strip().lower()
    if not fn:
        return False
    bare = os.path.basename(fn)
    # Always skip archives / compressed bundles
    for ext in SKIP_TRACK_EXTENSIONS:
        if bare.lower().endswith(ext):
            return True
    # Skip known non-track filenames
    if bare in SKIP_TRACK_FILENAMES:
        return True
    # Skip genome FASTA reference rows (TRACK_ID == 'dna')
    if label == 'dna':
        for ext in GENOME_FASTA_EXTENSIONS:
            if bare.lower().endswith(ext):
                return True
    return False


def get_common_name(base, organism):
    """Query organisms/{organism}/organism.sqlite for the common name."""
    db = os.path.join(base, 'organisms', organism, 'organism.sqlite')
    if not os.path.exists(db):
        return ''
    try:
        conn = sqlite3.connect(db)
        row = conn.execute('SELECT common_name FROM organism LIMIT 1').fetchone()
        conn.close()
        return row[0] if row and row[0] else ''
    except Exception:
        return ''


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────
def main():
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    meta_file = os.path.join(base, 'metadata.contents.txt')
    scp_file  = os.path.join(base, 'scp_commands.sh')
    out_file  = os.path.join(base, 'metadata', 'track_sheets_reformatted.xlsx')

    if not os.path.exists(meta_file):
        print(f"ERROR: {meta_file} not found"); sys.exit(1)
    if not os.path.exists(scp_file):
        print(f"ERROR: {scp_file} not found"); sys.exit(1)

    print("Loading organism groups …")
    groups_lookup = load_groups(base)
    print(f"  {len(groups_lookup)} entries in organism_assembly_groups.json")

    # Parse full tracks server file listing for TRACK_PATH fallback
    tracks_listing = os.path.join(base, 'tracks_files.txt')
    print("Parsing tracks_files.txt …")
    tracks_files_map = parse_tracks_files(tracks_listing)
    print(f"  {sum(len(v) for v in tracks_files_map.values())} file entries across {len(tracks_files_map)} (organism, assembly) pairs")

    # Parse scp_evidence.sh for evidence file URLs (with correct geneset paths)
    # and the geneset_map used to fix scp_commands.sh-derived URLs
    scp_ev_file = os.path.join(base, 'scp_evidence.sh')
    print("Parsing scp_evidence.sh …")
    evidence_by_code, geneset_map = parse_scp_evidence(scp_ev_file)
    supplement_geneset_map(geneset_map, os.path.join(base, 'move_gff_to_geneset.sh'))
    print(f"  {len(evidence_by_code)} organism codes with evidence files")
    print(f"  {len(geneset_map)} (organism, assembly) geneset entries")

    print("Parsing scp_commands.sh …")
    org_map, file_map = parse_scp_commands(scp_file)
    print(f"  {len(org_map)} organisms with valid new paths")

    # Parse genomes_find_results.txt for files located on the genomes server
    genomes_find_file = os.path.join(base, 'genomes_find_results.txt')
    print("Parsing genomes_find_results.txt …")
    genomes_map = parse_genomes_find_results(genomes_find_file, org_map, geneset_map)
    print(f"  {len(genomes_map)} filename→URL mappings from genomes find results")

    print("Parsing metadata.contents.txt …")
    metadata = parse_metadata_contents(meta_file)
    print(f"  {len(metadata)} organism code blocks found")

    # Suppress openpyxl's Excel 31-char sheet name warning (Google Sheets supports 100)
    import warnings
    warnings.filterwarnings('ignore', category=UserWarning,
                            message='Title is more than 31 characters')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    matched = 0
    skipped_pub = []
    skipped_no_meta = []
    used_tab_names = {}   # tab_name -> count, for dedup
    toc_entries = []      # (tab_name, organism, assembly, n_rows)

    # Codes to skip entirely (superseded assemblies — use the newer version instead)
    SKIP_CODES = {'PVA1_v1'}  # PVA1 superseded by PVA2 (different assembly)

    # --- Sheets for organisms in scp_map (skip _pub codes) ---
    for code in sorted(org_map.keys()):
        # Skip public/published variants and explicitly superseded codes
        if '_pub' in code.lower() or code.lower().endswith('_pub') or code in SKIP_CODES:
            skipped_pub.append(code)
            continue

        info = org_map[code]
        organism = info['organism']
        assembly = info['assembly']

        # Tab name: Organism_Assembly (up to 100 chars for Google Sheets)
        tab_name = f"{organism}_{assembly}"[:100]

        # Deduplicate tab names (shouldn't happen after skipping _pub, but just in case)
        if tab_name in used_tab_names:
            used_tab_names[tab_name] += 1
            tab_name = f"{tab_name[:95]}_{used_tab_names[tab_name]}"
        else:
            used_tab_names[tab_name] = 1

        rows = metadata.get(code, [])
        fp_map = file_map.get(code, {})

        if not rows and not fp_map:
            skipped_no_meta.append(code)
            continue

        # short_code strips version suffix: 'ACA1_v1' → 'ACA1'
        short_code = re.sub(r'_v\d+$', '', code, flags=re.IGNORECASE)
        ev_map    = evidence_by_code.get(short_code, {})
        geneset   = geneset_map.get((organism, assembly), '')
        trk_map   = tracks_files_map.get((organism, assembly), {})

        ws = wb.create_sheet(title=tab_name)

        # Header row
        ws.append(NEW_HEADERS)
        style_header_row(ws, len(NEW_HEADERS))
        ws.freeze_panes = 'A2'

        # Metadata rows (converted from old format)
        # The first 3 non-empty rows may be auto-registered tracks (DNA/genemodels/
        # transcriptmodels) — skip those only when they appear in the leading position.
        leading_skips = 0
        for old_row in rows:
            if all(v == '' for v in old_row.values()):
                continue
            if should_skip_row(old_row):
                continue
            if leading_skips < 3:
                label = old_row.get('label', '').strip().lower()
                if (label == 'dna'
                        or label.endswith('_genemodels')
                        or label.endswith('_transcriptmodels')):
                    leading_skips += 1
                    continue
                # Non-auto row encountered — stop checking leading rows
                leading_skips = 3
            new_row = convert_row(old_row, fp_map, evidence_map=ev_map,
                                  geneset=geneset, tracks_map=trk_map,
                                  genomes_map=genomes_map)
            ws.append([new_row.get(col, '') for col in NEW_HEADERS])

        # Any files in scp_map that have NO metadata row yet
        # (i.e., their filename doesn't appear in any row's FILENAME)
        used_files = set()
        for old_row in rows:
            fn = old_row.get('filename', '').strip()
            if fn:
                used_files.add(os.path.basename(fn))

        for fn, dest in sorted(fp_map.items()):
            bare = os.path.basename(fn)
            if bare in used_files:
                continue
            # Stub row with just FILENAME + TRACK_PATH (insert geneset if known)
            if dest and geneset:
                dest = re.sub(r'/gff(/|$)', f'/{geneset}/gff\\1', dest, count=1)
            new_row = {col: '' for col in NEW_HEADERS}
            new_row['FILENAME']   = bare
            new_row['TRACK_PATH'] = dest
            ws.append([new_row.get(col, '') for col in NEW_HEADERS])

        auto_width(ws, NEW_HEADERS)

        # Style body rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT

        common_name = get_common_name(base, organism)
        groups_str  = groups_lookup.get((organism, assembly), '')
        toc_entries.append((tab_name, organism, assembly,
                            ws.max_row - 1, common_name, groups_str))
        matched += 1
        print(f"  ✓ {tab_name}  ({len(rows)} metadata rows, {len(fp_map)} scp files)")

    # ── Add Nematostella vectensis sheet from its pre-updated TSV ──
    # Check multiple candidate locations
    for _candidate in [
        os.path.join(base, 'Nematostella_vectensis_updated.tsv'),
        os.path.join(base, 'organisms', 'Amphimedon_queenslandica',
                     'GCF_000090795.2', 'metazoa_r62', 'Nematostella_vectensis_updated.tsv'),
    ]:
        if os.path.exists(_candidate):
            nvec_tsv = _candidate
            break
    else:
        nvec_tsv = os.path.join(base, 'Nematostella_vectensis_updated.tsv')  # will trigger warning
    if os.path.exists(nvec_tsv):
        nvec_tab  = 'Nematostella_vectensis_GCA_033964005.1'
        nvec_base = TRACKS_BASE_URL + 'Nematostella_vectensis/'

        # Column name mapping for Nvec TSV → new headers
        NVEC_MAP = {
            'label':               'TRACK_ID',
            'key':                 'NAME',
            'technique':           'technique',
            'category':            'CATEGORY',
            'institute':           'institute',
            'source':              'source',
            'experiment':          'experiment',
            'developmental-stage': 'developmental-stage',
            'tissue':              'tissue',
            'condition':           'condition',
            'summary':             'summary',
            'citation':            'citation',
            'project':             'project',
            'accession':           'accession',
            'date':                'date',
            'analyst':             'analyst',
            'filename':            'FILENAME',
            'sciprj':              'SCIPRJ',
            'access_level':        'ACCESS',
            'biosample':           'biosample',
            'ngs_file':            'NGS_file',
            'mlong':               'MLONG',
        }

        ws_nvec = wb.create_sheet(title=nvec_tab)
        ws_nvec.append(NEW_HEADERS)
        style_header_row(ws_nvec, len(NEW_HEADERS))
        ws_nvec.freeze_panes = 'A2'

        with open(nvec_tsv, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.DictReader(f, delimiter='\t')
            nvec_rows = 0
            for old_row in reader:
                # Use label/filename columns to build a minimal dict for skip check
                _skip_check = {
                    'label':    old_row.get('label', ''),
                    'filename': old_row.get('filename', ''),
                }
                if should_skip_row(_skip_check):
                    continue

                new = {col: '' for col in NEW_HEADERS}
                for old_col, val in old_row.items():
                    mapped = NVEC_MAP.get(old_col.strip().lower())
                    if mapped and val:
                        if mapped == 'ACCESS':
                            new[mapped] = normalize_access(val)
                        else:
                            new[mapped] = val.strip()

                # Derive TRACK_PATH from filename
                fn = new.get('FILENAME', '').strip()
                if fn and ('/' in fn or fn.startswith('GCA_')):
                    new['TRACK_PATH'] = nvec_base + fn.lstrip('/')
                elif fn:
                    # Bare filename — try tracks_files_map then genomes_map
                    _nvec_trk = tracks_files_map.get(
                        ('Nematostella_vectensis', 'GCA_033964005.1'), {})
                    _bare = os.path.basename(fn)
                    _dest = _nvec_trk.get(fn, '') or _nvec_trk.get(_bare, '')
                    if not _dest and genomes_map:
                        _dest = genomes_map.get(fn, '') or genomes_map.get(_bare, '')
                    if _dest:
                        new['TRACK_PATH'] = _dest

                ws_nvec.append([new.get(col, '') for col in NEW_HEADERS])
                nvec_rows += 1

        auto_width(ws_nvec, NEW_HEADERS)
        for row in ws_nvec.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT

        nvec_common = get_common_name(base, 'Nematostella_vectensis')
        nvec_groups = groups_lookup.get(('Nematostella_vectensis', 'GCA_033964005.1'), '')
        toc_entries.append((nvec_tab, 'Nematostella_vectensis', 'GCA_033964005.1',
                            nvec_rows, nvec_common, nvec_groups))
        matched += 1
        print(f"  ✓ {nvec_tab}  ({nvec_rows} rows from updated TSV)")
    else:
        print(f"  WARNING: {nvec_tsv} not found — Nematostella not included")

    # ── Build Table of Contents as the FIRST sheet (after all organism sheets collected) ──
    toc = wb.create_sheet(title="Table of Contents", index=0)

    toc.append(["Tab", "Organism", "Common Name", "Groups", "Assembly", "Track rows"])
    hdr_fill = PatternFill("solid", fgColor="0891B2")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in toc[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
    toc.freeze_panes = 'A2'
    toc.column_dimensions['A'].width = 55
    toc.column_dimensions['B'].width = 28
    toc.column_dimensions['C'].width = 24
    toc.column_dimensions['D'].width = 20
    toc.column_dimensions['E'].width = 22
    toc.column_dimensions['F'].width = 12

    link_font     = Font(color="0891B2", underline="single", size=10)
    body_font_toc = Font(size=10)

    for row_i, (tab_name, organism, assembly, n_rows, common_name, groups_str) in enumerate(toc_entries, start=2):
        cell = toc.cell(row=row_i, column=1, value=tab_name)
        cell.hyperlink = f"#{tab_name}!A1"
        cell.font = link_font
        toc.cell(row=row_i, column=2, value=organism.replace('_', ' ')).font = body_font_toc
        toc.cell(row=row_i, column=3, value=common_name).font = body_font_toc
        toc.cell(row=row_i, column=4, value=groups_str).font = body_font_toc
        toc.cell(row=row_i, column=5, value=assembly).font = body_font_toc
        toc.cell(row=row_i, column=6, value=n_rows).font = body_font_toc

    print(f"\n✓ Table of Contents written ({len(toc_entries)} entries)")
    print(f"\nCreated {matched} organism sheets + 1 TOC = {matched + 1} total")
    if skipped_pub:
        print(f"Skipped (_pub): {', '.join(skipped_pub)}")
    if skipped_no_meta:
        print(f"Skipped (no metadata + no scp files): {', '.join(skipped_no_meta)}")

    wb.save(out_file)
    print(f"\nSaved → {out_file}")
    return out_file


if __name__ == '__main__':
    main()
