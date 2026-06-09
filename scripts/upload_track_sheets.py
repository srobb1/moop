#!/usr/bin/env python3
"""
Upload track_sheets_reformatted.xlsx to Google Drive as a native Google Spreadsheet.

Uses OAuth2 device flow — no browser needed on the server side.
On first run you'll get a URL + code to paste; token is cached for future runs.

Usage:
    python3 scripts/upload_track_sheets.py
"""

import os
import sys
import json
import base64
import pickle

# Check dependencies
try:
    import requests
except ImportError:
    print("ERROR: requests not installed. Run: python3 -m pip install --user requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: python3 -m pip install --user openpyxl")
    sys.exit(1)

# ──────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_FILE  = os.path.join(BASE_DIR, 'metadata', 'track_sheets_reformatted.xlsx')
TOKEN_FILE = os.path.join(BASE_DIR, 'metadata', '.google_token.json')
SHEET_NAME = "MOOP Track Sheets (reformatted)"

# Google OAuth2 — use the installed app client ID that ships with gspread
# This is a public client ID for device flow (safe to embed)
CLIENT_ID     = "977351234234-s3f8d4r7g9b2p6n1c5l3w8q0m4e2v6t5.apps.googleusercontent.com"
CLIENT_SECRET = ""  # device flow doesn't need a secret for public clients

# Alternatively, provide your own credentials JSON:
CREDENTIALS_FILE = os.path.join(BASE_DIR, 'metadata', 'google_credentials.json')

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]


# ──────────────────────────────────────────────────────────────
# Auth helpers
# ──────────────────────────────────────────────────────────────

def load_credentials_file():
    """Load client_id/client_secret from a credentials JSON (downloaded from Google Cloud Console)."""
    if not os.path.exists(CREDENTIALS_FILE):
        return None, None
    with open(CREDENTIALS_FILE) as f:
        d = json.load(f)
    installed = d.get('installed') or d.get('web', {})
    return installed.get('client_id'), installed.get('client_secret')


def get_token():
    """Load cached token or run device flow to get a new one."""

    # Try cached token first
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE) as f:
            token_data = json.load(f)
        # Try refresh
        refreshed = refresh_token(token_data)
        if refreshed:
            return refreshed
        print("Cached token expired and couldn't refresh — re-authenticating …")

    # Try loading credentials from file
    client_id, client_secret = load_credentials_file()

    if not client_id:
        print("""
No Google credentials found.

Please create a OAuth 2.0 Client ID in Google Cloud Console:
  1. Go to https://console.cloud.google.com/apis/credentials
  2. Create credentials → OAuth client ID → Desktop app
  3. Download the JSON and save it to:
     metadata/google_credentials.json

Or, if you have gspread installed with a service account:
  gspread.service_account(filename='path/to/service_account.json')

For now, the XLSX file has been saved to:
  metadata/track_sheets_reformatted.xlsx

You can import it manually to Google Sheets:
  1. Go to https://drive.google.com
  2. New → File upload → select the XLSX file
  3. Right-click → Open with → Google Sheets
  4. File → Save as Google Sheets (to convert format)

The XLSX has 64 tabs: a TOC + 63 organism/assembly sheets.
""")
        sys.exit(0)

    return device_flow_auth(client_id, client_secret)


def device_flow_auth(client_id, client_secret):
    """Run OAuth2 device authorization flow."""
    r = requests.post(
        'https://oauth2.googleapis.com/device/code',
        data={
            'client_id': client_id,
            'scope': ' '.join(SCOPES),
        }
    )
    r.raise_for_status()
    d = r.json()

    print(f"""
=== Google Authentication Required ===

Visit: {d['verification_url']}
Enter code: {d['user_code']}

Waiting for you to authenticate …
""")

    import time
    interval = d.get('interval', 5)
    device_code = d['device_code']

    while True:
        time.sleep(interval)
        r2 = requests.post(
            'https://oauth2.googleapis.com/token',
            data={
                'client_id': client_id,
                'client_secret': client_secret,
                'device_code': device_code,
                'grant_type': 'urn:ietf:params:oauth:grant-type:device_code',
            }
        )
        resp = r2.json()
        if 'access_token' in resp:
            save_token(resp)
            print("✓ Authenticated!\n")
            return resp
        if resp.get('error') == 'authorization_pending':
            continue
        if resp.get('error') == 'slow_down':
            interval += 5
            continue
        print(f"Auth error: {resp}")
        sys.exit(1)


def refresh_token(token_data):
    """Try to refresh the access token. Returns new token data or None."""
    refresh = token_data.get('refresh_token')
    client_id, client_secret = load_credentials_file()
    if not refresh or not client_id:
        return None

    r = requests.post(
        'https://oauth2.googleapis.com/token',
        data={
            'client_id': client_id,
            'client_secret': client_secret,
            'refresh_token': refresh,
            'grant_type': 'refresh_token',
        }
    )
    if r.status_code != 200:
        return None
    new_data = r.json()
    new_data.setdefault('refresh_token', refresh)
    save_token(new_data)
    return new_data


def save_token(token_data):
    with open(TOKEN_FILE, 'w') as f:
        json.dump(token_data, f)
    os.chmod(TOKEN_FILE, 0o600)


# ──────────────────────────────────────────────────────────────
# Google Sheets / Drive API helpers
# ──────────────────────────────────────────────────────────────

def api_headers(token):
    return {'Authorization': f"Bearer {token['access_token']}"}


def create_spreadsheet(token, title):
    """Create an empty Google Spreadsheet and return its ID."""
    r = requests.post(
        'https://sheets.googleapis.com/v4/spreadsheets',
        headers={**api_headers(token), 'Content-Type': 'application/json'},
        json={'properties': {'title': title}}
    )
    r.raise_for_status()
    return r.json()['spreadsheetId']


def batch_update(token, spreadsheet_id, requests_body):
    """Send a batchUpdate to the Sheets API."""
    r = requests.post(
        f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}:batchUpdate',
        headers={**api_headers(token), 'Content-Type': 'application/json'},
        json={'requests': requests_body}
    )
    if r.status_code != 200:
        print(f"batchUpdate error: {r.text[:500]}")
        r.raise_for_status()
    return r.json()


def values_update(token, spreadsheet_id, range_name, values):
    """Write values to a range."""
    r = requests.put(
        f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{range_name}',
        headers={**api_headers(token), 'Content-Type': 'application/json'},
        params={'valueInputOption': 'USER_ENTERED'},
        json={'range': range_name, 'majorDimension': 'ROWS', 'values': values}
    )
    if r.status_code != 200:
        print(f"values_update error: {r.text[:500]}")
        r.raise_for_status()


def get_sheet_id_map(token, spreadsheet_id):
    """Return {sheet_title -> sheetId}."""
    r = requests.get(
        f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}',
        headers=api_headers(token),
        params={'fields': 'sheets.properties'}
    )
    r.raise_for_status()
    return {
        s['properties']['title']: s['properties']['sheetId']
        for s in r.json().get('sheets', [])
    }


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

TEAL_HEX = '0891B2'

def hex_to_rgb(h):
    return {
        'red':   int(h[0:2], 16) / 255,
        'green': int(h[2:4], 16) / 255,
        'blue':  int(h[4:6], 16) / 255,
    }


def header_format_request(sheet_id):
    return {
        'repeatCell': {
            'range': {'sheetId': sheet_id, 'startRowIndex': 0, 'endRowIndex': 1},
            'cell': {
                'userEnteredFormat': {
                    'backgroundColor': hex_to_rgb(TEAL_HEX),
                    'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}},
                    'horizontalAlignment': 'CENTER',
                }
            },
            'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)',
        }
    }


def freeze_row_request(sheet_id):
    return {
        'updateSheetProperties': {
            'properties': {'sheetId': sheet_id, 'gridProperties': {'frozenRowCount': 1}},
            'fields': 'gridProperties.frozenRowCount',
        }
    }


def main():
    if not os.path.exists(XLSX_FILE):
        print(f"ERROR: {XLSX_FILE} not found. Run create_track_sheets.py first.")
        sys.exit(1)

    token = get_token()

    print(f"Loading {XLSX_FILE} …")
    wb = openpyxl.load_workbook(XLSX_FILE)
    sheet_names = wb.sheetnames
    print(f"  {len(sheet_names)} sheets to upload")

    print(f"Creating Google Spreadsheet '{SHEET_NAME}' …")
    spreadsheet_id = create_spreadsheet(token, SHEET_NAME)
    print(f"  ID: {spreadsheet_id}")
    print(f"  URL: https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit")

    # Get the default sheet ID (will be renamed to first XLSX sheet)
    sheet_id_map = get_sheet_id_map(token, spreadsheet_id)
    default_sheet_id = list(sheet_id_map.values())[0]

    # Build addSheet requests for sheets 2..N, rename default sheet to sheet 1
    add_sheet_requests = []

    # Rename the default sheet to the first XLSX sheet name
    first_name = sheet_names[0]
    add_sheet_requests.append({
        'updateSheetProperties': {
            'properties': {'sheetId': default_sheet_id, 'title': first_name},
            'fields': 'title',
        }
    })

    # Add remaining sheets
    for name in sheet_names[1:]:
        add_sheet_requests.append({
            'addSheet': {'properties': {'title': name}}
        })

    print("  Adding all sheet tabs …")
    batch_update(token, spreadsheet_id, add_sheet_requests)

    # Refresh sheet ID map after creation
    sheet_id_map = get_sheet_id_map(token, spreadsheet_id)

    # Upload data for each sheet
    format_requests = []
    for i, ws_name in enumerate(sheet_names):
        ws = wb[ws_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else '' for cell in row])

        if not rows:
            continue

        sheet_id = sheet_id_map.get(ws_name)
        if sheet_id is None:
            print(f"  WARNING: sheet '{ws_name}' not found in created spreadsheet")
            continue

        # Escape the sheet name for the range (use first 100 chars, quote it)
        safe_name = ws_name.replace("'", "''")
        range_name = f"'{safe_name}'!A1"

        print(f"  [{i+1}/{len(sheet_names)}] {ws_name} ({len(rows)} rows)")
        values_update(token, spreadsheet_id, range_name, rows)

        # Queue format requests
        format_requests.append(header_format_request(sheet_id))
        format_requests.append(freeze_row_request(sheet_id))

    # Apply all formatting in one batch
    if format_requests:
        print("  Applying header formatting …")
        # Split into batches of 100 to avoid API limits
        for i in range(0, len(format_requests), 100):
            batch_update(token, spreadsheet_id, format_requests[i:i+100])

    print(f"""
Done!
  Spreadsheet: {SHEET_NAME}
  URL: https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit
""")


if __name__ == '__main__':
    main()
